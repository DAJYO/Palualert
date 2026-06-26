"""
Lecture et validation d'un fichier de mise à jour des données.

Format attendu (identique au panel) :
    district, date, temp_moy, temp_max, humidite, precip_mensuel, cas

Utilisé par l'onglet « Charger un fichier ». Distingue les **erreurs**
bloquantes (intégration impossible) des **avertissements** non bloquants.
"""

from io import BytesIO

import pandas as pd

COLONNES_ATTENDUES = [
    "district", "date", "temp_moy", "temp_max", "humidite", "precip_mensuel", "cas",
]
COLONNES_NUM = ["temp_moy", "temp_max", "humidite", "precip_mensuel", "cas"]

INSTRUCTIONS = [
    "MODELE DE SAISIE — Surveillance Palustre (Extreme-Nord)",
    "",
    "Une ligne = un district pour un mois donne.",
    "Ne PAS modifier les noms de district ni l'ordre des colonnes.",
    "",
    "Colonnes :",
    "  district        : nom exact du district (deja pre-rempli).",
    "  date            : 1er du mois au format AAAA-MM-JJ (ex. 2026-07-01).",
    "  temp_moy        : temperature moyenne du mois.",
    "  temp_max        : temperature maximale du mois.",
    "  humidite        : humidite en % (valeur entre 0 et 100).",
    "  precip_mensuel  : precipitations du mois (>= 0).",
    "  cas             : nombre de cas de paludisme (>= 0, entier).",
    "",
    "Regles de validation (onglet « Charger un fichier ») :",
    "  - districts inconnus, dates illisibles, humidite hors 0-100 et",
    "    valeurs negatives -> BLOQUANT.",
    "  - cases vides, dates hors 1er du mois, doublons -> avertissement.",
    "",
    "Pour plusieurs mois : dupliquer les 33 lignes en changeant la date.",
]


def _date_par_defaut():
    """Premier jour du mois suivant (ex. 2026-07-01)."""
    t = pd.Timestamp.today().normalize().replace(day=1) + pd.DateOffset(months=1)
    return t.strftime("%Y-%m-%d")


def _table_modele(districts, date):
    df = pd.DataFrame({"district": list(districts)})
    df["date"] = date
    for c in COLONNES_NUM:
        df[c] = ""
    return df[COLONNES_ATTENDUES]


def modele_csv(districts, date=None):
    """Octets d'un modèle CSV (33 districts pré-remplis, valeurs à compléter)."""
    date = date or _date_par_defaut()
    return _table_modele(districts, date).to_csv(index=False).encode("utf-8")


def modele_xlsx(districts, date=None):
    """Octets d'un modèle Excel à deux feuilles : Instructions + Donnees."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    date = date or _date_par_defaut()
    wb = Workbook()

    # Feuille Instructions
    ws_i = wb.active
    ws_i.title = "Instructions"
    for i, ligne in enumerate(INSTRUCTIONS, start=1):
        c = ws_i.cell(row=i, column=1, value=ligne)
        if i == 1:
            c.font = Font(bold=True, size=13)
    ws_i.column_dimensions["A"].width = 75

    # Feuille Donnees
    ws_d = wb.create_sheet("Donnees")
    entete = Font(bold=True, color="FFFFFF")
    fond = PatternFill("solid", fgColor="1F6FB2")
    for j, nom in enumerate(COLONNES_ATTENDUES, start=1):
        c = ws_d.cell(row=1, column=j, value=nom)
        c.font = entete
        c.fill = fond
        c.alignment = Alignment(horizontal="center")
    for i, d in enumerate(districts, start=2):
        ws_d.cell(row=i, column=1, value=d)
        ws_d.cell(row=i, column=2, value=date)
    for col in "ABCDEFG":
        ws_d.column_dimensions[col].width = 16
    ws_d.freeze_panes = "A2"

    flux = BytesIO()
    wb.save(flux)
    return flux.getvalue()


def lire_et_valider(fichier, districts_valides):
    """Lit un fichier Excel/CSV et vérifie sa conformité au format du panel.

    Paramètres
    ----------
    fichier : objet `st.file_uploader` ou chemin.
    districts_valides : liste des noms de districts attendus.

    Retour
    ------
    (df, erreurs, avertissements)
      - df : DataFrame nettoyé si lisible, sinon None ;
      - erreurs : problèmes bloquants ;
      - avertissements : anomalies non bloquantes.
    """
    erreurs, avertissements = [], []

    # 1) Lecture du fichier
    try:
        nom = getattr(fichier, "name", str(fichier)).lower()
        if nom.endswith(".csv"):
            df = pd.read_csv(fichier)
        else:
            # Excel : privilégie une feuille « Donnees » si elle existe
            xls = pd.ExcelFile(fichier)
            feuille = next(
                (s for s in xls.sheet_names if s.strip().lower() in ("donnees", "données")),
                0,
            )
            df = pd.read_excel(xls, sheet_name=feuille)
    except Exception as e:  # lecture impossible : on s'arrête là
        return None, [f"Impossible de lire le fichier : {e}"], []

    # 2) Colonnes
    manquantes = [c for c in COLONNES_ATTENDUES if c not in df.columns]
    if manquantes:
        return None, [f"Colonnes manquantes : {', '.join(manquantes)}"], []
    df = df[COLONNES_ATTENDUES].copy()  # ignore les colonnes en trop

    # 3) Districts
    inconnus = sorted(set(df["district"].dropna().astype(str)) - set(districts_valides))
    if inconnus:
        erreurs.append("Districts non reconnus : " + ", ".join(inconnus))

    # 4) Dates (AAAA-MM-JJ, 1er du mois)
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    n_invalides = int(df["date"].isna().sum())
    if n_invalides:
        erreurs.append(
            f"{n_invalides} date(s) illisible(s) (format attendu : AAAA-MM-JJ)."
        )
    pas_1er = int(df["date"].dropna().dt.day.ne(1).sum())
    if pas_1er:
        avertissements.append(f"{pas_1er} date(s) ne tombent pas le 1er du mois.")

    # 5) Valeurs numériques
    for c in COLONNES_NUM:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    n_vides = int(df[COLONNES_NUM].isna().sum().sum())
    if n_vides:
        avertissements.append(
            f"{n_vides} valeur(s) numérique(s) vide(s) ou non numérique(s)."
        )
    hum = df["humidite"].dropna()
    if ((hum < 0) | (hum > 100)).any():
        erreurs.append("Humidité hors bornes (doit être comprise entre 0 et 100 %).")
    for c in ["precip_mensuel", "cas"]:
        if (df[c].dropna() < 0).any():
            erreurs.append(f"Valeurs négatives détectées dans « {c} ».")

    # 6) Doublons (district, date)
    dup = int(df.duplicated(subset=["district", "date"]).sum())
    if dup:
        avertissements.append(
            f"{dup} ligne(s) en double (même district et même mois)."
        )

    return df, erreurs, avertissements
